"""Excel and CSV export.

Header fields and line items are 1:N, so they get separate sheets joined on
doc_id. Putting both in one flat sheet forces either header duplication on
every line or ragged blanks.

Cell typing is not cosmetic: written as a general cell, invoice number 0012345
becomes 12345 and PO reference 1-2 becomes 2-Jan. That is silent data loss, and
it is the most common complaint about extraction tools.
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import logging
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import DocResult, DocStatus, Status

log = logging.getLogger(__name__)

TEXT_FMT = "@"
MONEY_FMT = "#,##0.00"
DATE_FMT = "DD-MMM-YYYY"          # unambiguous, unlike DD/MM/YYYY

HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
AMBER = PatternFill("solid", fgColor="FFF1C2")
RED = PatternFill("solid", fgColor="FFD4D4")
BLUE = PatternFill("solid", fgColor="D6E7FF")     # human-corrected

# Fields that must never be written as numbers.
FORCE_TEXT = {"invoice_number", "po_number", "gst_number", "gstin", "pan",
              "hsn", "sac", "item_code", "customer_gstin", "vendor_gstin",
              "sr_no", "reference", "challan_number", "eway_bill"}

EXCEL_MAX_ROWS = 1_048_576


def _is_money(name: str, value) -> bool:
    return isinstance(value, Decimal) and name.lower() not in FORCE_TEXT


def _write_cell(ws, row: int, col: int, name: str, value, *,
                status: Status | None = None, confidence: float | None = None,
                reasons=None, corrected: bool = False):
    cell = ws.cell(row=row, column=col)
    key = (name or "").lower()

    if value is None or value == "":
        cell.value = ""
    elif key in FORCE_TEXT:
        cell.value = str(value)
        cell.number_format = TEXT_FMT
    elif isinstance(value, dt.date):
        cell.value = value
        cell.number_format = DATE_FMT
    elif isinstance(value, Decimal):
        cell.value = float(value)
        cell.number_format = MONEY_FMT
    elif isinstance(value, (int, float)):
        cell.value = value
    else:
        cell.value = str(value)
        if str(value)[:1].isdigit() and any(c in str(value) for c in "-/"):
            cell.number_format = TEXT_FMT     # stop Excel making it a date

    if corrected:
        cell.fill = BLUE
    elif status in (Status.AMBIGUOUS, Status.INVALID, Status.NOT_FOUND,
                    Status.FAILED):
        cell.fill = RED
    elif status == Status.LOW_CONFIDENCE:
        cell.fill = AMBER

    if status and status not in (Status.EXTRACTED, Status.MISSING):
        note = f"{status.value}"
        if confidence is not None:
            note += f"\nconfidence: {confidence:.2f}"
        if reasons:
            note += "\n" + ", ".join(reasons[:4])
        cell.comment = Comment(note, "DocumentExtractor")
    return cell


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if ncols:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws, max_width: int = 44) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = min(max_width,
                                      max(widths.get(cell.column, 10),
                                          len(str(cell.value)) + 2))
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def collect_export_fields(results: list[DocResult], templates: list) -> list[str]:
    """The field names for the second, clean export - the union of
    column_order across whichever templates were actually matched in this
    batch, not every template that happened to be loaded. A template's
    column_order is exactly the fields whose checkbox was ticked when it was
    built (see the template editor), so this needs no separate bookkeeping
    of "what did the user select" - the template already records it.
    """
    used_ids = {r.template_id for r in results if r.template_id}
    by_id = {t.template_id: t for t in templates}
    seen: set[str] = set()
    ordered: list[str] = []
    for tid in used_ids:
        tpl = by_id.get(tid)
        if tpl is None:
            continue
        for name in tpl.output.column_order:
            if name not in seen:
                seen.add(name)
                ordered.append(name)
    return ordered


def build_fields_only_workbook(results: list[DocResult], selected_fields: list[str]
                               ) -> Workbook:
    """The second, clean export: only the fields the user actually checked
    in the review screen, one row per document, no status columns, no
    confidence notes, no colour coding. This is the "just give me the data"
    file - build_workbook's Documents sheet is the audit trail; this is the
    deliverable someone hands to a bookkeeper or loads into another system
    without needing to understand what NEEDS_REVIEW or BELOW_FLOOR_80 mean.

    selected_fields controls both which columns appear and their order -
    if a document's own template does not define one of them, that cell is
    simply blank, the same convention build_workbook already uses so a
    batch spanning several vendors still produces one flat table.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Fields"

    headers = ["doc_id", "source_file"] + selected_fields
    ws.append(headers)
    _style_header(ws, len(headers))

    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r.doc_id)
        ws.cell(row=row_idx, column=2, value=r.source_name)
        for offset, name in enumerate(selected_fields):
            fr = r.fields.get(name)
            # No status/confidence/colour here on purpose - this sheet
            # answers "what is the value", not "how sure are we". Anyone
            # who needs that goes to the full report instead.
            _write_cell(ws, row_idx, 3 + offset, name, fr.value if fr else None)
    _autosize(ws)
    return wb


def build_workbook(results: list[DocResult], manifest: dict,
                   *, include_field_detail: bool = False) -> Workbook:
    wb = Workbook()

    # ---- Summary --------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    counts = {s.value: 0 for s in DocStatus}
    for r in results:
        counts[r.status.value] = counts.get(r.status.value, 0) + 1

    rows = [
        ("Run ID", manifest.get("run_id", "")),
        ("Started (UTC)", manifest.get("started_utc", "")),
        ("Finished (UTC)", manifest.get("finished_utc", "")),
        ("Application version", manifest.get("app_version", "")),
        ("Build", manifest.get("build", "")),
        ("Input folder", manifest.get("input_dir", "")),
        ("", ""),
        ("Documents processed", len(results)),
        ("Clean (OK)", counts.get("OK", 0)),
        ("Needs review", counts.get("NEEDS_REVIEW", 0)),
        ("Failed", counts.get("FAILED", 0)),
        ("Total line items", sum(len(r.line_items) for r in results)),
        ("", ""),
        ("Confidence floor", manifest.get("confidence_floor", "")),
        ("Template min score", manifest.get("template_min_score", "")),
        ("Template margin", manifest.get("template_margin", "")),
        ("OCR enabled", manifest.get("ocr_enabled", False)),
    ]
    ws.cell(row=1, column=1, value="Setting").font = HEADER_FONT
    ws.cell(row=1, column=2, value="Value").font = HEADER_FONT
    _style_header(ws, 2)
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v if v is not None else "")

    start = len(rows) + 4
    ws.cell(row=start, column=1, value="Template").font = Font(bold=True)
    ws.cell(row=start, column=2, value="Documents").font = Font(bold=True)
    ws.cell(row=start, column=3, value="Needs review").font = Font(bold=True)
    per_template: dict[str, list[int]] = {}
    for r in results:
        entry = per_template.setdefault(r.template_id or "(unmatched)", [0, 0])
        entry[0] += 1
        if r.status != DocStatus.OK:
            entry[1] += 1
    for i, (tid, (total, flagged)) in enumerate(sorted(per_template.items()),
                                                start=start + 1):
        ws.cell(row=i, column=1, value=tid)
        ws.cell(row=i, column=2, value=total)
        ws.cell(row=i, column=3, value=flagged)
    _autosize(ws)

    # ---- Exceptions (second tab: the one users open first) ---------------
    ws = wb.create_sheet("Exceptions")
    headers = ["doc_id", "source_file", "field", "raw_value", "status",
               "reason_codes", "confidence", "page"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        for f in r.flagged_fields():
            ws.append([r.doc_id, r.source_name, f.name, f.raw, f.status.value,
                       ", ".join(f.reasons), round(f.confidence, 3), f.page])
        for c in r.cross_checks:
            if not c.passed and c.severity == "review":
                ws.append([r.doc_id, r.source_name, c.check_id, "",
                           "CROSS_CHECK_FAILED", c.detail, "", ""])
    _autosize(ws)

    # ---- Documents -------------------------------------------------------
    ws = wb.create_sheet("Documents")
    field_names: list[str] = []
    for r in results:
        for name in r.fields:
            if name not in field_names:
                field_names.append(name)
    order = manifest.get("column_order") or []
    field_names.sort(key=lambda n: (order.index(n) if n in order else 999, n))

    headers = (["doc_id", "source_file", "sha256", "pages", "template_id",
                "template_version", "vendor", "doc_type", "match_score",
                "status", "review_reasons"] + field_names +
               ["line_item_count", "processed_utc"])
    ws.append(headers)
    _style_header(ws, len(headers))

    for row_idx, r in enumerate(results, start=2):
        base = [r.doc_id, r.source_name, r.sha256[:16], r.page_count,
                r.template_id, r.template_version, r.vendor, r.doc_type,
                round(r.match_score, 3), r.status.value,
                ", ".join(r.all_reasons()[:6])]
        for col, value in enumerate(base, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        for offset, name in enumerate(field_names):
            fr = r.fields.get(name)
            _write_cell(ws, row_idx, len(base) + 1 + offset, name,
                        fr.value if fr else None,
                        status=fr.status if fr else None,
                        confidence=fr.confidence if fr else None,
                        reasons=fr.reasons if fr else None,
                        corrected=fr.corrected if fr else False)
        ws.cell(row=row_idx, column=len(base) + len(field_names) + 1,
                value=len(r.line_items))
        ws.cell(row=row_idx, column=len(base) + len(field_names) + 2,
                value=r.processed_utc)
    _autosize(ws)

    # ---- LineItems -------------------------------------------------------
    ws = wb.create_sheet("LineItems")
    col_names: list[str] = []
    for r in results:
        for li in r.line_items:
            for name in li.cells:
                if name not in col_names:
                    col_names.append(name)
    headers = ["doc_id", "source_file", "line_no"] + col_names + \
              ["page", "status", "reasons"]
    ws.append(headers)
    _style_header(ws, len(headers))

    row_idx = 2
    truncated = False
    for r in results:
        for li in r.line_items:
            if row_idx >= EXCEL_MAX_ROWS - 10:
                truncated = True
                break
            ws.cell(row=row_idx, column=1, value=r.doc_id)
            ws.cell(row=row_idx, column=2, value=r.source_name)
            ws.cell(row=row_idx, column=3, value=li.index)
            for offset, name in enumerate(col_names):
                _write_cell(ws, row_idx, 4 + offset, name, li.cells.get(name),
                            status=li.status, confidence=li.confidence,
                            reasons=li.reasons)
            ws.cell(row=row_idx, column=4 + len(col_names), value=li.page)
            ws.cell(row=row_idx, column=5 + len(col_names), value=li.status.value)
            ws.cell(row=row_idx, column=6 + len(col_names),
                    value=", ".join(li.reasons))
            row_idx += 1
        if truncated:
            break
    if truncated:
        log.warning("LineItems truncated at the Excel row limit")
    _autosize(ws)

    # ---- Failed ----------------------------------------------------------
    ws = wb.create_sheet("Failed")
    headers = ["source_file", "reason", "error", "sha256"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        if r.status == DocStatus.FAILED:
            ws.append([r.source_name, ", ".join(r.all_reasons()[:6]),
                       (r.error or "")[:250], r.sha256[:16]])
    _autosize(ws)

    # ---- FieldDetail (optional) -----------------------------------------
    if include_field_detail:
        ws = wb.create_sheet("FieldDetail")
        headers = ["doc_id", "source_file", "field", "value", "raw", "status",
                   "confidence", "strategy", "page", "reasons"]
        ws.append(headers)
        _style_header(ws, len(headers))
        for r in results:
            for name, f in r.fields.items():
                ws.append([r.doc_id, r.source_name, name,
                           "" if f.value is None else str(f.value), f.raw,
                           f.status.value, round(f.confidence, 3), f.strategy,
                           f.page, ", ".join(f.reasons)])
        _autosize(ws)

    return wb


def write_outputs(results: list[DocResult], manifest: dict, output_dir: Path,
                  *, include_field_detail: bool = False,
                  selected_fields: list[str] | None = None) -> dict:
    """Write the workbook plus CSV/JSONL sidecars. Returns the paths written.

    selected_fields is optional: pass the list of fields the user checked in
    the review screen to also get a second, clean workbook containing only
    those columns. Omit it (the default) and only the full report is
    produced - existing callers that have never heard of field selection
    keep working exactly as before.
    """
    output_dir = Path(output_dir)
    (output_dir / "data").mkdir(parents=True, exist_ok=True)

    run_id = manifest.get("run_id", "run")
    xlsx_path = output_dir / f"DocumentExtractor_{run_id}.xlsx"
    wb = build_workbook(results, manifest, include_field_detail=include_field_detail)
    wb.save(xlsx_path)

    # Excel is a report for humans. Anyone wiring this into an ERP wants these.
    docs_csv = output_dir / "data" / "documents.csv"
    field_names: list[str] = []
    for r in results:
        for n in r.fields:
            if n not in field_names:
                field_names.append(n)
    with docs_csv.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["doc_id", "source_file", "template_id", "vendor", "status"]
                   + field_names + ["line_item_count"])
        for r in results:
            w.writerow([r.doc_id, r.source_name, r.template_id, r.vendor,
                        r.status.value]
                       + ["" if (r.fields.get(n) is None or
                                 r.fields[n].value is None)
                          else str(r.fields[n].value) for n in field_names]
                       + [len(r.line_items)])

    lines_csv = output_dir / "data" / "line_items.csv"
    col_names: list[str] = []
    for r in results:
        for li in r.line_items:
            for n in li.cells:
                if n not in col_names:
                    col_names.append(n)
    with lines_csv.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["doc_id", "line_no"] + col_names + ["page", "status"])
        for r in results:
            for li in r.line_items:
                w.writerow([r.doc_id, li.index]
                           + ["" if li.cells.get(n) is None else str(li.cells[n])
                              for n in col_names]
                           + [li.page, li.status.value])

    jsonl = output_dir / "data" / "results.jsonl"
    with jsonl.open("w", encoding="utf-8") as fh:
        for r in results:
            fh.write(json.dumps(r.to_dict(), ensure_ascii=False,
                                default=str) + "\n")

    report = write_html_report(results, manifest, output_dir / "report.html")

    paths = {"xlsx": xlsx_path, "documents_csv": docs_csv,
            "line_items_csv": lines_csv, "jsonl": jsonl, "report": report}

    if selected_fields:
        fields_wb = build_fields_only_workbook(results, selected_fields)
        fields_xlsx = output_dir / f"DocumentExtractor_{run_id}_fields.xlsx"
        fields_wb.save(fields_xlsx)
        paths["fields_xlsx"] = fields_xlsx

    return paths


def write_html_report(results: list[DocResult], manifest: dict,
                      path: Path) -> Path:
    """A shareable processing report that does not require Excel."""
    counts: dict[str, int] = {}
    reasons: dict[str, int] = {}
    for r in results:
        counts[r.status.value] = counts.get(r.status.value, 0) + 1
        for reason in r.all_reasons():
            reasons[reason] = reasons.get(reason, 0) + 1

    rows = "".join(
        f"<tr class='{r.status.value.lower()}'><td>{_esc(r.source_name)}</td>"
        f"<td>{_esc(r.template_id)}</td><td>{r.status.value}</td>"
        f"<td>{len(r.line_items)}</td>"
        f"<td>{_esc(', '.join(r.all_reasons()[:4]))}</td></tr>"
        for r in results)
    reason_rows = "".join(
        f"<tr><td>{_esc(k)}</td><td>{v}</td></tr>"
        for k, v in sorted(reasons.items(), key=lambda kv: -kv[1]))

    html = f"""<!doctype html><html><head><meta charset="utf-8">
<title>Processing report {_esc(manifest.get('run_id',''))}</title>
<style>
body{{font-family:Segoe UI,system-ui,sans-serif;margin:2rem;color:#1a1a1a}}
h1{{font-size:1.4rem}} h2{{font-size:1.05rem;margin-top:2rem}}
table{{border-collapse:collapse;width:100%;font-size:.85rem;margin-top:.5rem}}
th,td{{border:1px solid #ddd;padding:.4rem .55rem;text-align:left}}
th{{background:#1F3B57;color:#fff}}
tr.needs_review{{background:#FFF1C2}} tr.failed{{background:#FFD4D4}}
.cards{{display:flex;gap:1rem;margin:1rem 0}}
.card{{border:1px solid #ddd;border-radius:6px;padding:.8rem 1.2rem;min-width:7rem}}
.card b{{display:block;font-size:1.6rem}}
</style></head><body>
<h1>Processing report — {_esc(manifest.get('run_id',''))}</h1>
<p>{_esc(manifest.get('started_utc',''))} → {_esc(manifest.get('finished_utc',''))}</p>
<div class="cards">
  <div class="card"><b>{len(results)}</b>documents</div>
  <div class="card"><b>{counts.get('OK',0)}</b>clean</div>
  <div class="card"><b>{counts.get('NEEDS_REVIEW',0)}</b>need review</div>
  <div class="card"><b>{counts.get('FAILED',0)}</b>failed</div>
</div>
<h2>Reason codes</h2><table><tr><th>Reason</th><th>Count</th></tr>{reason_rows}</table>
<h2>Documents</h2><table>
<tr><th>File</th><th>Template</th><th>Status</th><th>Lines</th><th>Reasons</th></tr>
{rows}</table></body></html>"""
    Path(path).write_text(html, encoding="utf-8")
    return Path(path)


def _esc(s) -> str:
    return (str(s or "").replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))
